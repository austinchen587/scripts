# file_processor.py
import os
import docx
import PyPDF2
import pandas as pd
from PIL import Image
import logging
from pathlib import Path
import pdfplumber
import zipfile
import base64
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

logger = logging.getLogger(__name__)

class ExcelImageExtractor:
    """Excel图片提取工具类"""
    @staticmethod
    def extract_images(file_path, output_dir="temp_images"):
        """提取Excel中的图片并返回Base64编码列表"""
        images = []
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
        
        try:
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for img in sheet._images:
                    if isinstance(img, OpenpyxlImage):
                        # 保存图片到临时文件
                        img_path = output_dir / f"{sheet_name}_img_{len(images)}.png"
                        with open(img_path, "wb") as f:
                            f.write(img._data())
                        
                        # 转换为Base64
                        with open(img_path, "rb") as f:
                            images.append({
                                "sheet": sheet_name,
                                "base64": base64.b64encode(f.read()).decode("utf-8"),
                                "format": "png",
                                "path": str(img_path)
                            })
            return images
        except Exception as e:
            logger.error(f"Excel图片提取失败: {e}")
            return []

class FileProcessor:
    def __init__(self):
        self.supported_text_extensions = ['.txt', '.docx', '.pdf', '.doc', '.xls', '.xlsx']
        self.supported_image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
    
    def extract_text_from_file(self, file_path):
        """从文件中提取文本内容"""
        file_path = Path(file_path)
        if not file_path.exists():
            logger.error(f"文件不存在: {file_path}")
            return ""
        
        extension = file_path.suffix.lower()
        
        try:
            if extension == '.txt':
                text = self._read_txt(file_path)
                logger.info(f"提取到TXT文件内容: {text[:100]}...")
                return text
            elif extension == '.docx':
                text = self._read_docx(file_path)
                logger.info(f"提取到DOCX文件内容: {text[:100]}...")
                return text
            elif extension == '.pdf':
                text = self._read_pdf(file_path)
                logger.info(f"提取到PDF文件内容: {text[:100]}...")
                return text
            elif extension == '.doc':
                text = self._read_doc(file_path)
                logger.info(f"提取到DOC文件内容: {text[:100]}...")
                return text
            elif extension in ['.xls', '.xlsx']:
                text = self._read_excel(file_path)
                logger.info(f"提取到Excel文件内容: {text[:100]}...")
                return text
            else:
                logger.warning(f"不支持的文件格式: {extension}")
                return f"[不支持的文件格式: {extension}]"
        except Exception as e:
            logger.error(f"文件读取失败 {file_path}: {e}")
            return f"[文件读取失败: {e}]"
    
    def _read_txt(self, file_path):
        """读取txt文件"""
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    
    def _read_docx(self, file_path):
        """读取docx文件"""
        doc = docx.Document(file_path)
        return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
    
    def _read_doc(self, file_path):
        """读取doc文件"""
        try:
            return f"[DOC文件格式，建议转换为DOCX查看内容: {file_path}]"
        except Exception as e:
            return f"[DOC文件读取失败: {e}]"
    
    def _read_pdf(self, file_path):
        """读取pdf文件"""
        text = ""
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + '\n'
            return text.strip() if text else "[PDF无文本内容或为扫描件]"
        except Exception as e:
            logger.error(f"PDF读取失败 {file_path}: {e}")
            try:
                with open(file_path, 'rb') as f:
                    pdf_reader = PyPDF2.PdfReader(f)
                    for page in pdf_reader.pages:
                        text += page.extract_text() + '\n'
                return text.strip() if text else "[PDF内容提取困难]"
            except Exception as e2:
                return f"[PDF读取彻底失败: {e2}]"
    
    def _read_excel(self, file_path):
        """读取Excel文件（多工作表支持）"""
        try:
            # 根据文件扩展名自动选择引擎
            engine = 'xlrd' if str(file_path).endswith('.xls') else 'openpyxl'
                
            content_summary = []
            content_summary.append(f"Excel文件: {Path(file_path).name}")
            
            # 获取所有工作表数据
            with pd.ExcelFile(file_path, engine=engine) as xls:
                sheet_names = xls.sheet_names
                content_summary.append(f"工作表数量: {len(sheet_names)}")
                
                for sheet_name in sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    content_summary.append(f"\n=== 工作表 '{sheet_name}' ===")
                    content_summary.append(f"行数: {len(df)}, 列数: {len(df.columns)}")
                    
                    # 处理DISPIMG公式
                    img_cols = [col for col in df.columns if df[col].astype(str).str.contains('DISPIMG').any()]
                    if img_cols:
                        content_summary.append(f"包含图片的列: {', '.join(img_cols)}")
                    
                    content_summary.append("列名: " + ", ".join(df.columns.tolist()))
                    
                    if len(df) > 0:
                        content_summary.append("\n数据样例:")
                        sample_rows = min(3, len(df))
                        for i in range(sample_rows):
                            row_data = []
                            for col in df.columns:
                                value = df.iloc[i][col]
                                if pd.notna(value):
                                    row_data.append(f"{col}: {value}")
                            content_summary.append(f"行{i+1}: {' | '.join(row_data)}")
            
            return '\n'.join(content_summary)
        except ImportError as e:
            logger.error(f"缺少依赖库: {str(e)}")
            if "xlrd" in str(e):
                return "[错误] 请安装xlrd: pip install xlrd==2.0.1"
            elif "openpyxl" in str(e):
                return "[错误] 请安装openpyxl: pip install openpyxl"
        except Exception as e:
            logger.error(f"Excel读取失败 {file_path}: {e}")
            return f"[Excel表格数据，读取错误: {e}]"
    
    def extract_file_structured_data(self, file_path):
        """提取文件的结构化数据（支持多工作表+图片）"""
        extension = Path(file_path).suffix.lower()
        
        if extension in ['.xls', '.xlsx']:
            return self._extract_excel_data(file_path)
        elif extension == '.pdf':
            return self._extract_pdf_data(file_path)
        else:
            return {"file_type": extension, "content": self.extract_text_from_file(file_path)}
    
    def _extract_excel_data(self, file_path):
        """提取Excel结构化数据（增强版）"""
        try:
            engine = 'xlrd' if str(file_path).endswith('.xls') else 'openpyxl'
            structured_data = {
                "file_type": "excel",
                "file_name": Path(file_path).name,
                "sheets": [],
                "images": []  # 新增图片信息字段
            }

            # 提取图片（仅支持.xlsx）
            if str(file_path).endswith('.xlsx'):
                structured_data["images"] = ExcelImageExtractor.extract_images(file_path)

            # 读取所有工作表
            with pd.ExcelFile(file_path, engine=engine) as xls:
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    
                    # 标记包含DISPIMG的列
                    img_columns = [
                        col for col in df.columns 
                        if df[col].astype(str).str.contains('DISPIMG').any()
                    ]
                    
                    sheet_data = {
                        "sheet_name": sheet_name,
                        "columns": df.columns.tolist(),
                        "row_count": len(df),
                        "image_columns": img_columns,
                        "data": df.where(pd.notna(df), None).to_dict(orient="records")
                    }
                    structured_data["sheets"].append(sheet_data)

            return structured_data
        except Exception as e:
            logger.error(f"Excel结构化数据提取失败: {e}")
            return {"file_type": "excel", "error": str(e)}
    
    def _extract_pdf_data(self, file_path):
        """提取PDF结构化数据"""
        text_content = self._read_pdf(file_path)
        return {
            "file_type": "pdf",
            "file_name": Path(file_path).name,
            "content": text_content,
            "key_info": self._extract_key_info_from_text(text_content)
        }
    
    def _extract_key_info_from_text(self, text):
        """从文本中提取关键信息"""
        key_phrases = []
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if any(keyword in line.lower() for keyword in 
                  ['型号', '规格', '品牌', '数量', '单位', '要求', '参数', '技术']):
                key_phrases.append(line)
        return key_phrases[:10]
    
    def is_image_file(self, file_path):
        """判断是否为图片文件"""
        return Path(file_path).suffix.lower() in self.supported_image_extensions
    
    def get_file_info(self, file_path):
        """获取文件信息"""
        file_path = Path(file_path)
        return {
            'path': str(file_path),
            'exists': file_path.exists(),
            'is_image': self.is_image_file(file_path),
            'size': file_path.stat().st_size if file_path.exists() else 0,
            'extension': file_path.suffix.lower()
        }
